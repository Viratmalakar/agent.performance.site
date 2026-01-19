from flask import Flask, render_template, request, redirect, url_for, session, send_file
import pandas as pd
import io
from datetime import datetime

app = Flask(__name__)
app.secret_key = "agentdashboard"

# ---------------- UTIL ----------------
def fix_time(x):
    if pd.isna(x) or str(x).strip().lower() in ["nan","-",""]:
        return "00:00:00"
    return str(x)

def tsec(t):
    try:
        h,m,s=map(int,str(t).split(":"))
        return h*3600+m*60+s
    except:
        return 0

def stime(s):
    h=s//3600
    m=(s%3600)//60
    s=s%60
    return f"{h:02d}:{m:02d}:{s:02d}"

# ---------------- ROUTES ----------------

@app.route("/")
def upload():
    return render_template("upload.html")

@app.route("/process",methods=["POST"])
def process():

    agent=pd.read_excel(request.files["agent"],engine="openpyxl")
    cdr=pd.read_excel(request.files["cdr"],engine="openpyxl")

    agent=agent.fillna("00:00:00")

    emp=agent.columns[1]
    full=agent.columns[2]
    login=agent.columns[3]
    talk=agent.columns[5]

    t=agent.columns[19]; u=agent.columns[20]
    w=agent.columns[22]; x=agent.columns[23]; y=agent.columns[24]

    c_emp=cdr.columns[1]
    camp=cdr.columns[6]
    disp=cdr.columns[25]

    cdr[camp]=cdr[camp].astype(str)
    cdr[disp]=cdr[disp].astype(str)

    mature=cdr[cdr[disp].str.contains("callmature|transfer",case=False,na=False)]
    ib=mature[mature[camp].str.upper()=="CSRINBOUND"]

    mature_cnt=mature.groupby(c_emp).size()
    ib_cnt=ib.groupby(c_emp).size()

    final=pd.DataFrame()

    final["Agent Name"]=agent[emp]
    final["Agent Full Name"]=agent[full]

    final["Total Login Time"]=agent[login].apply(fix_time)

    final["Total Break"]=(agent[t].apply(tsec)+agent[w].apply(tsec)+agent[y].apply(tsec)).apply(stime)
    final["Total Meeting"]=(agent[u].apply(tsec)+agent[x].apply(tsec)).apply(stime)
    final["Total Net Login"]=(agent[login].apply(tsec)-final["Total Break"].apply(tsec)).apply(stime)

    final["Total Talk Time"]=agent[talk].apply(fix_time)

    final["Total Call"]=final["Agent Name"].map(mature_cnt).fillna(0).astype(int)
    final["IB Mature"]=final["Agent Name"].map(ib_cnt).fillna(0).astype(int)
    final["OB Mature"]=final["Total Call"]-final["IB Mature"]

    final["AHT"]=(final["Total Talk Time"].apply(tsec)/final["Total Call"].replace(0,1)).astype(int).apply(stime)

    # ---- PERMANENT JUNK ROW DELETE ----
    final=final[final["Agent Name"].astype(str).str.lower()!="nan"]
    final=final[final["Agent Name"]!="Agent Name"]

    final=final.reset_index(drop=True)

    # ---- COLUMN SEQUENCE ----
    final=final[[
        "Agent Name","Agent Full Name","Total Login Time","Total Net Login",
        "Total Break","Total Meeting","AHT","Total Call","IB Mature","OB Mature"
    ]]

    # ---- GRAND TOTAL ----
    gt={}
    gt["TOTAL IVR HIT"]=int(cdr[cdr[camp].str.upper()=="CSRINBOUND"].shape[0])
    gt["TOTAL MATURE"]=int(final["Total Call"].sum())
    gt["IB MATURE"]=int(final["IB Mature"].sum())
    gt["OB MATURE"]=int(final["OB Mature"].sum())
    gt["TOTAL TALK TIME"]=stime(final["Total Login Time"].apply(tsec).sum())
    gt["AHT"]=stime(int(final["Total Login Time"].apply(tsec).sum()/max(1,gt["TOTAL MATURE"])))
    gt["LOGIN COUNT"]=int(final["Agent Name"].nunique())

    session["data"]=final.to_dict(orient="records")
    session["gt"]=gt

    return redirect(url_for("result"))

@app.route("/result")
def result():
    return render_template("result.html",
        data=session.get("data",[]),
        gt=session.get("gt",{})
    )

# ---------------- EXPORT ----------------

@app.route("/export")
def export():
    data=pd.DataFrame(session.get("data",[]))

    out=io.BytesIO()
    with pd.ExcelWriter(out,engine="openpyxl") as writer:
        data.to_excel(writer,index=False,sheet_name="Report")
        ws=writer.sheets["Report"]

        from openpyxl.styles import Font,PatternFill,Border,Side

        header_fill=PatternFill("solid",fgColor="1fa463")
        header_font=Font(color="FFFFFF",bold=True)
        border=Border(left=Side(style="thin"),right=Side(style="thin"),
                      top=Side(style="thin"),bottom=Side(style="thin"))

        for c in ws[1]:
            c.fill=header_fill
            c.font=header_font
            c.border=border

        for r in ws.iter_rows(min_row=2):
            for c in r:
                c.border=border

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=20

    out.seek(0)
    now=datetime.now().strftime("%d-%m-%y %H-%M-%S")
    fname=f"Agent_Performance_Report_Chandan-Malakar_{now}.xlsx"
    return send_file(out,download_name=fname,as_attachment=True)

# ---------------- RUN ----------------
if __name__=="__main__":
    app.run(debug=True)
